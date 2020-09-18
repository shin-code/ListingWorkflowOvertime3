import sys
import time
import os
from selenium import webdriver
from selenium.webdriver.support.select import Select
from bs4 import BeautifulSoup
import openpyxl as px
from datetime import datetime
import myFn

# 定数定義
XLSX_FILE = "ListWorkflowOvertime3.xlsx"
SHEET_NAME = "Overtime3"

# パラメータ
dnId = input("ログインID >> ")
dnPw = input("パスワード >> ")
fromDate = myFn.text_to_date(input("抽出範囲日付FROM ( yyyy/mm/dd形式 ) >> "))  # 日付変換

'''
**************************************************
 出力先エクセルを開く
**************************************************
'''
# 出力先ファイル存在する場合はファイルを開く
if os.path.isfile(XLSX_FILE):

   wb = px.load_workbook(XLSX_FILE)
   ws = wb[SHEET_NAME]


else:  # 出力先ファイル存在しない場合はファイルを新規作成

   wb = px.Workbook()
   ws = wb.active
   ws.title = SHEET_NAME  # シート名

   # 作成情報
   ws['A1'].value = '時間外申請一覧（システム３課）'

   # タイトル行
   ws['A2'].value = '申請者'
   ws['B2'].value = '申請日時'
   ws['C2'].value = '日付'
   ws['D2'].value = '氏名'
   ws['E2'].value = '実績開始時間'
   ws['F2'].value = '実績終了時間'
   ws['G2'].value = '朝のサーバーチェック'
   ws['H2'].value = '申請時間'
   ws['I2'].value = '申請深夜時間'
   ws['J2'].value = '一覧作成日時'

   wb.save(XLSX_FILE)  # 一旦保存

'''
**************************************************
 desknet'sログイン → ワークフロー遷移
**************************************************
'''
# chromeからdesknet'sを開く
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS  # pylint: disable=no-member
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)

driver = webdriver.Chrome(resource_path('./chromedriver.exe'))

driver.get("https://dkn.e-omc.jp/cgi-bin/dneo/dneo.cgi?")

# id
id = driver.find_element_by_name("UserID")
id.send_keys(dnId)

# pw
pw = driver.find_element_by_name("_word")
pw.send_keys(dnPw)

# ログイン
id.submit()

time.sleep(3)

# ワークフロー
driver.get("https://dkn.e-omc.jp/cgi-bin/dneo/zflow.cgi?cmd=flowindex")

time.sleep(3)

# ワークフローステータス -> 完了
wfStatusElem = driver.find_element_by_id("flow-list-type-sel")  # element取得
wfStatusObj = Select(wfStatusElem)
wfStatusObj.select_by_visible_text("完了")

time.sleep(1)

wfTblThElem = driver.find_elements_by_class_name("flow-list-line")  # 明細行elements取得

'''
**************************************************
 ワークフロー１件ずつ処理 → 一覧化対象idを作成
**************************************************
'''
listingIds = []  # 一覧化対象id

# 明細行ループ
for i in range(len(wfTblThElem)):

   wfTblTdElem = wfTblThElem[i].find_elements_by_tag_name("td")

   # 状況が"完了"以外はスキップ
   if wfTblTdElem[3].text != "完了":
      continue

   # 表題が"時間外申請書"以外はスキップ
   if wfTblTdElem[4].text != "時間外申請書":
      continue

   # 作成日付を取得
   slashCount = wfTblTdElem[6].text.count("/")  # スラッシュの出現回数
   if slashCount == 1:  # 年が省略されているので今年を付加
      textDate = str(datetime.now().year) + "/" + wfTblTdElem[6].text[0:5]
      textTime = wfTblTdElem[6].text[6:]
   else:
      textDate = wfTblTdElem[6].text[0:10]
      textTime = wfTblTdElem[6].text[11:]
   wfMakeAplydate = myFn.text_to_date(textDate)  # 日付変換

   # 作成日付が抽出範囲日付FROM以下の場合、ループを抜ける
   if wfMakeAplydate < fromDate:
      break

   # 作成日付（yyyy/mm/dd HH:MM)
   wfMakeAplydate = wfMakeAplydate.strftime('%Y/%m/%d') + " " + textTime

   # 申請者
   wfMakeAplycant = wfTblTdElem[5].text
   wfMakeAplycant = wfMakeAplycant[0:wfMakeAplycant.find(" ")]  # 宮崎一郎 [システム○課]のスペース前部分を抽出

   # すでにエクセルに存在するかチェック
   flgXlsExist = False
   for j in range(1, ws.max_row+1):
      # 申請者と申請日時で検索
      if ws.cell(row=j, column=1).value == wfMakeAplycant \
         and ws.cell(row=j, column=2).value == wfMakeAplydate:
         flgXlsExist = True
         break
   
   # すでにエクセルに存在した場合はスキップ
   if flgXlsExist:
      continue

   # チェックボックスのidを取得
   wfTblThChkElem = wfTblThElem[i].find_elements_by_class_name("co-chk")  # チェックボックスelement取得
   wfTblThChkInputElem = wfTblThChkElem[0].find_element_by_name("id")  # チェックボックス配下のinput取得
   wfId = wfTblThChkInputElem.get_attribute("value")

   # 一覧化対象idを作成
   listingId = {
      'id': wfId,
      'wfMakeAplycant' : wfMakeAplycant,
      'wfMakeAplydate' : wfMakeAplydate,
      }
   listingIds.append(listingId)

'''
**************************************************
 一覧化対象idを1件ずつ処理
**************************************************
'''
wsRow = ws.max_row + 1  # エクセル書込行開始行（最終行取得+1）

# 一覧下対象idをループ
for i in range(len(listingIds)):

   # 取得したidで単票表示
   driver.get("https://dkn.e-omc.jp/cgi-bin/dneo/zflow.cgi?cmd=flowindex#cmd=flowdisp&id=" + listingIds[i]['id'])
   time.sleep(1)

   """
   明細部分の基本的な構成

   氏名        宮崎 一郎    ←1つのパーツ form-parts
    ↑          ↑
   fontタグ    fontタグ
   """
   wfFormParts = driver.find_elements_by_class_name("form-parts")

   # 日付
   wfFpFonts = wfFormParts[2].find_elements_by_tag_name("font")
   wfDate = wfFpFonts[1].text

   # 氏名
   wfFpFonts = wfFormParts[3].find_elements_by_tag_name("font")
   wfName = wfFpFonts[1].text

   # 実績開始時間
   wfFpFonts = wfFormParts[15].find_elements_by_tag_name("font")
   wfStartTime = wfFpFonts[0].text

   # 実績終了時刻
   wfFpFonts = wfFormParts[16].find_elements_by_tag_name("font")
   wfEndTime = wfFpFonts[1].text

   # 朝のサーバーチェック
   wfFpImg = wfFormParts[17].find_element_by_tag_name("img")
   if "form_checkbox_on.gif" in wfFpImg.get_attribute("src"):
      wfSvChk = 'True'
   else:
      wfSvChk = 'False'

   # 申告時間
   wfFpFonts = wfFormParts[18].find_elements_by_tag_name("font")
   wfOvertime = wfFpFonts[1].text

   # 深刻深夜時間
   wfFpFonts = wfFormParts[19].find_elements_by_tag_name("font")
   wfOvetimeMidnight = wfFpFonts[1].text

   # エクセル出力   
   ws.cell(row=wsRow, column=2).value  = listingIds[i]['wfMakeAplydate']  # 申請日時
   ws.cell(row=wsRow, column=1).value  = listingIds[i]['wfMakeAplycant']  # 申請者
   ws.cell(row=wsRow, column=3).value  = wfDate                           # 日付
   ws.cell(row=wsRow, column=4).value  = wfName                           # 氏名
   ws.cell(row=wsRow, column=5).value  = wfStartTime                      # 実績開始時間
   ws.cell(row=wsRow, column=6).value  = wfEndTime                        # 実績終了時間
   ws.cell(row=wsRow, column=7).value  = wfSvChk                          # 朝のサーバーチェック
   ws.cell(row=wsRow, column=8).value  = wfOvertime                       # 申請時間
   ws.cell(row=wsRow, column=9).value  = wfOvetimeMidnight                # 申請深夜時間
   ws.cell(row=wsRow, column=10).value = datetime.now()                   # 一覧作成日時

   wsRow += 1  # エクセル書き込み行カウントアップ

wb.save(XLSX_FILE)  # エクセル保存

driver.close()
driver.quit()